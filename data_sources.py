"""
data_sources.py
Builds normalized "evidence bundles" for stocks, either from local demo JSON
files (offline) or live from yfinance (NSE tickers, *.NS).

The evidence bundle schema (per stock) is the single source of truth that the
scoring engines (deterministic + LLM) are allowed to cite numbers from. Any
field that could not be computed is set to null AND its name is appended to
data_gaps, so downstream consumers can say "data unavailable" honestly instead
of inventing a figure.
"""

import json
import os
import statistics
import time
from datetime import datetime, timezone, timedelta

IST = timezone(timedelta(hours=5, minutes=30))

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEMO_DIR = os.path.join(BASE_DIR, "demo_data")
UNIVERSE_PATH = os.path.join(BASE_DIR, "universe.json")


def now_ist_str():
    return datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S IST")


def load_universe():
    """Load the editable ticker universe grouped by cap segment."""
    if not os.path.exists(UNIVERSE_PATH):
        return {"large": [], "mid": [], "small": []}
    with open(UNIVERSE_PATH, "r") as f:
        return json.load(f)


# --------------------------------------------------------------------------
# DEMO MODE
# --------------------------------------------------------------------------

def load_demo_bundles():
    """Load every pre-built evidence bundle in demo_data/*.json."""
    bundles = []
    if not os.path.isdir(DEMO_DIR):
        return bundles
    for fname in sorted(os.listdir(DEMO_DIR)):
        if fname.endswith(".json"):
            with open(os.path.join(DEMO_DIR, fname), "r") as f:
                try:
                    bundles.append(json.load(f))
                except json.JSONDecodeError:
                    continue
    return bundles


# --------------------------------------------------------------------------
# LIVE MODE (yfinance)
# --------------------------------------------------------------------------

def _safe(fn, *args, **kwargs):
    try:
        return fn(*args, **kwargs)
    except Exception:
        return None


def _pct(a, b):
    """Percent change of a relative to b, safely."""
    try:
        if a is None or b is None or b == 0:
            return None
        return round((a - b) / b * 100, 2)
    except Exception:
        return None


def build_evidence_bundle(symbol, name, cap_segment, sector=None):
    """
    Pull ~1 month daily OHLC + .info + .news for `symbol` (NSE ticker, e.g.
    RELIANCE.NS) via yfinance, and normalize into the evidence bundle schema.
    Never raises — any failure is recorded in data_gaps with nulled fields.
    """
    import yfinance as yf

    gaps = []
    bundle = {
        "symbol": symbol,
        "name": name or symbol.replace(".NS", ""),
        "cap_segment": cap_segment,
        "sector": sector,
        "price": {},
        "range_52w": {},
        "technicals": {},
        "analyst": {},
        "news": {},
        "data_gaps": gaps,
        "note": "Feed has no raw fundamental ratios (P/E, ROE, etc.) - not available in this pipeline.",
    }

    ticker = _safe(yf.Ticker, symbol)
    if ticker is None:
        gaps.append("ticker_unavailable")
        return bundle

    # Use a short fresh history request for the market data path. `Ticker.info`
    # can contain stale cached quote fields, so it is only a fallback for fields
    # that are not available from the latest history/fast_info.
    hist = _safe(lambda: ticker.history(period="1mo", interval="1d", auto_adjust=False))
    fast = _safe(lambda: ticker.fast_info) or {}
    info = _safe(lambda: ticker.info) or {}
    news = _safe(lambda: ticker.news) or []

    # ---- price block ----
    live = fast.get("last_price") or info.get("currentPrice") or info.get("regularMarketPrice")
    day_open = fast.get("open") or info.get("open") or info.get("regularMarketOpen")
    day_high = fast.get("day_high") or info.get("dayHigh") or info.get("regularMarketDayHigh")
    day_low = fast.get("day_low") or info.get("dayLow") or info.get("regularMarketDayLow")
    prev_close = fast.get("previous_close") or info.get("previousClose") or info.get("regularMarketPreviousClose")
    volume = info.get("volume") or info.get("regularMarketVolume")

    # The latest OHLC row is the authoritative fallback for the most recent
    # trading session, especially when quote metadata is stale.
    if hist is not None and not hist.empty:
        latest = hist.iloc[-1]
        if latest.get("Close") is not None:
            live = float(latest["Close"]) if live is None else live
        if day_open is None and latest.get("Open") is not None:
            day_open = float(latest["Open"])
        if day_high is None and latest.get("High") is not None:
            day_high = float(latest["High"])
        if day_low is None and latest.get("Low") is not None:
            day_low = float(latest["Low"])
        if volume is None and latest.get("Volume") is not None:
            volume = float(latest["Volume"])

    if prev_close is None and hist is not None and len(hist) >= 2:
        prev_close = float(hist["Close"].iloc[-2])
    if prev_close is None and hist is not None and len(hist) >= 2:
        prev_close = float(hist["Close"].iloc[-2])
    if volume is None and hist is not None and not hist.empty:
        volume = float(hist["Volume"].iloc[-1])

    day_change_pct = _pct(live, prev_close)
    for field, val in [("live", live), ("day_open", day_open), ("day_high", day_high),
                        ("day_low", day_low), ("prev_close", prev_close), ("volume", volume)]:
        if val is None:
            gaps.append(f"price.{field}")
    if day_change_pct is None:
        gaps.append("price.day_change_pct")

    bundle["price"] = {
        "live": live, "day_open": day_open, "day_high": day_high, "day_low": day_low,
        "prev_close": prev_close, "day_change_pct": day_change_pct, "volume": volume,
    }

    # ---- 52w range ----
    hi52 = info.get("fiftyTwoWeekHigh")
    lo52 = info.get("fiftyTwoWeekLow")
    pct_from_high = _pct(live, hi52) if (live and hi52) else None
    position_pct = None
    if live is not None and hi52 is not None and lo52 is not None and hi52 != lo52:
        position_pct = round((live - lo52) / (hi52 - lo52) * 100, 1)
    if hi52 is None:
        gaps.append("range_52w.high")
    if lo52 is None:
        gaps.append("range_52w.low")
    if position_pct is None:
        gaps.append("range_52w.position_pct")

    bundle["range_52w"] = {
        "high": hi52, "low": lo52, "pct_from_high": pct_from_high, "position_pct": position_pct,
    }

    # ---- technicals ----
    rvol = None
    sma_n = 20
    price_vs_sma_pct = None
    window_return_pct = None
    swing_high = None
    swing_low = None
    day_range_position_pct = None
    trend = None

    if hist is not None and not hist.empty:
        closes = hist["Close"].dropna().tolist()
        vols = hist["Volume"].dropna().tolist()
        if len(vols) >= 2:
            avg_prior_vol = statistics.mean(vols[:-1]) if len(vols) > 1 else None
            today_vol = vols[-1]
            if avg_prior_vol and avg_prior_vol > 0:
                rvol = round(today_vol / avg_prior_vol, 2)
        if len(closes) >= 2:
            window_return_pct = _pct(closes[-1], closes[0])
        if len(closes) >= sma_n:
            sma = statistics.mean(closes[-sma_n:])
            price_vs_sma_pct = _pct(closes[-1], sma)
        elif len(closes) >= 5:
            sma = statistics.mean(closes)
            price_vs_sma_pct = _pct(closes[-1], sma)
        swing_high = max(closes) if closes else None
        swing_low = min(closes) if closes else None
        if price_vs_sma_pct is not None:
            if price_vs_sma_pct > 1:
                trend = "up"
            elif price_vs_sma_pct < -1:
                trend = "down"
            else:
                trend = "sideways"
    else:
        gaps.append("technicals.history_unavailable")

    if day_high is not None and day_low is not None and live is not None and day_high != day_low:
        day_range_position_pct = round((live - day_low) / (day_high - day_low) * 100, 1)

    for field, val in [("rvol", rvol), ("price_vs_sma_pct", price_vs_sma_pct),
                        ("window_return_pct", window_return_pct), ("swing_high", swing_high),
                        ("swing_low", swing_low), ("day_range_position_pct", day_range_position_pct),
                        ("trend", trend)]:
        if val is None:
            gaps.append(f"technicals.{field}")

    bundle["technicals"] = {
        "rvol": rvol, "price_vs_sma_pct": price_vs_sma_pct, "window_return_pct": window_return_pct,
        "swing_high": swing_high, "swing_low": swing_low,
        "day_range_position_pct": day_range_position_pct, "trend": trend,
    }

    # ---- analyst ----
    target_mean = info.get("targetMeanPrice")
    target_low = info.get("targetLowPrice")
    target_high = info.get("targetHighPrice")
    num_analysts = info.get("numberOfAnalystOpinions")
    consensus = info.get("recommendationKey")
    upside_pct = _pct(target_mean, live) if (target_mean and live) else None

    rec_map = {"strong_buy": (95, 3, 2), "buy": (80, 15, 5), "hold": (40, 40, 20),
               "underperform": (15, 60, 25), "sell": (5, 70, 25)}
    buy_pct = hold_pct = sell_pct = None
    if consensus in rec_map:
        buy_pct, hold_pct, sell_pct = rec_map[consensus]

    for field, val in [("target_mean", target_mean), ("num_analysts", num_analysts),
                        ("consensus", consensus), ("upside_pct", upside_pct)]:
        if val is None:
            gaps.append(f"analyst.{field}")

    bundle["analyst"] = {
        "consensus": consensus, "num_analysts": num_analysts,
        "buy_pct": buy_pct, "hold_pct": hold_pct, "sell_pct": sell_pct,
        "target_mean": target_mean, "target_low": target_low, "target_high": target_high,
        "upside_pct": upside_pct,
    }

    # ---- news ----
    recent = []
    pos = neg = neu = 0
    POS_WORDS = ["surge", "beat", "upgrade", "growth", "profit", "record", "rally",
                 "gain", "positive", "strong", "outperform", "buy"]
    NEG_WORDS = ["fall", "miss", "downgrade", "loss", "decline", "probe", "fraud",
                 "weak", "negative", "sell", "cut", "lawsuit"]
    for item in (news or [])[:8]:
        title = (item.get("title") or item.get("content", {}).get("title", "")) if isinstance(item, dict) else ""
        if not title and isinstance(item, dict):
            content = item.get("content") or {}
            title = content.get("title", "")
        if not title:
            continue
        low = title.lower()
        tone = "neutral"
        if any(w in low for w in POS_WORDS):
            tone = "positive"
            pos += 1
        elif any(w in low for w in NEG_WORDS):
            tone = "negative"
            neg += 1
        else:
            neu += 1
        recent.append({"title": title, "tone": tone})

    if not recent:
        gaps.append("news.recent")

    bundle["news"] = {
        "total": len(recent), "positive": pos, "negative": neg, "neutral": neu,
        "recent": recent[:5],
    }

    return bundle


# --------------------------------------------------------------------------
# EXCEL "FINAL LIST" IMPORT
# --------------------------------------------------------------------------

def parse_excel_final_list(file_bytes, preferred_sheet="Final List"):
    """
    Parse an uploaded .xlsx workbook and pull out a ticker list.

    Looks for a sheet named `preferred_sheet` (case-insensitive substring
    match, e.g. "Final List"); if not found, falls back to the active sheet.
    Within that sheet, finds the header row containing an "NSE Code" column
    (case-insensitive) and reads every non-empty row below it until a blank
    row, collecting the ticker symbol (and, if present, a CAR / verdict
    column for display only - it is not fed to the agents as ground truth).

    Returns: {"sheet": <name>, "tickers": [{"symbol": "XXX.NS", "name": "XXX",
              "car": "<CAR value or None>"}], "count": N}
    Raises ValueError with a human-readable message if nothing usable is found.
    """
    import io
    import openpyxl

    # Not read_only: some workbooks (e.g. ones hand-edited or exported from
    # other tools) have missing/stale dimension metadata, which makes
    # read_only mode report max_row/max_column as None and breaks iteration.
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    sheet_name = None
    for name in wb.sheetnames:
        if preferred_sheet.lower() in name.lower():
            sheet_name = name
            break
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    all_rows = list(ws.iter_rows(values_only=False))

    # find the header row: the first row containing a cell that looks like "NSE Code"
    header_row_idx = None
    header_cells = None
    for scan_i, row in enumerate(all_rows[:40]):
        values = [c.value for c in row]
        for i, v in enumerate(values):
            if isinstance(v, str) and "nse code" in v.strip().lower():
                header_row_idx = scan_i
                header_cells = values
                break
        if header_row_idx is not None:
            break

    if header_row_idx is None:
        raise ValueError(
            f"Could not find an 'NSE Code' column in sheet '{sheet_name}'. "
            "Make sure the workbook has a column named 'NSE Code'."
        )

    col_map = {}
    for i, v in enumerate(header_cells):
        if not isinstance(v, str):
            continue
        key = v.strip().lower()
        if "nse code" in key:
            col_map["symbol"] = i
        elif key == "car":
            col_map["car"] = i
        elif "cmp" in key:
            col_map["cmp"] = i

    tickers = []
    seen = set()
    for row in all_rows[header_row_idx + 1:]:
        values = [c.value for c in row]
        if col_map["symbol"] >= len(values):
            continue
        raw_symbol = values[col_map["symbol"]]
        if raw_symbol is None or not str(raw_symbol).strip():
            continue
        code = str(raw_symbol).strip().upper()
        if code in ("NSE CODE", "STOCKS IN BULL RUN"):
            continue
        symbol = code if code.endswith(".NS") else f"{code}.NS"
        if symbol in seen:
            continue
        seen.add(symbol)
        car = values[col_map["car"]] if "car" in col_map and col_map["car"] < len(values) else None
        tickers.append({"symbol": symbol, "name": code, "car": car})

    if not tickers:
        raise ValueError(f"Sheet '{sheet_name}' has an 'NSE Code' header but no ticker rows under it.")

    return {"sheet": sheet_name, "tickers": tickers, "count": len(tickers)}


def build_bundles_for_tickers(tickers, cap_segment="watchlist"):
    """
    Build one evidence bundle per ticker dict ({"symbol","name",...}), all
    tagged with `cap_segment`. Used for an already-curated list (e.g. an
    uploaded "Final List") where no further screening/shortlisting is needed
    - every ticker the user gave us goes straight to the debate stage.
    """
    bundles = []
    for t in tickers:
        symbol = t["symbol"] if isinstance(t, dict) else t
        name = t.get("name") if isinstance(t, dict) else None
        bundles.append(build_evidence_bundle(symbol, name, cap_segment))
    return bundles


def screen_and_shortlist(universe, shortlist_per_bucket=4, sector_map=None):
    """
    For each cap bucket (large/mid/small), build evidence bundles for every
    ticker, then keep the top `shortlist_per_bucket` by absolute day-change %.
    Returns (scanned_bundles, shortlisted_bundles).
    """
    sector_map = sector_map or {}
    all_bundles = []
    for segment, tickers in universe.items():
        for entry in tickers:
            if isinstance(entry, dict):
                symbol = entry.get("symbol")
                name = entry.get("name")
            else:
                symbol, name = entry, None
            sector = sector_map.get(symbol)
            b = build_evidence_bundle(symbol, name, segment, sector)
            all_bundles.append(b)

    shortlisted = []
    for segment in universe.keys():
        seg_bundles = [b for b in all_bundles if b["cap_segment"] == segment]
        seg_bundles.sort(
            key=lambda b: abs(b["price"].get("day_change_pct") or 0), reverse=True
        )
        shortlisted.extend(seg_bundles[:shortlist_per_bucket])

    return all_bundles, shortlisted


def parse_google_sheet_rows(payload, preferred_sheet="Final List"):
    """Parse Google Visualization API table JSON into the same ticker structure as Excel."""
    cols = payload.get("table", {}).get("cols", [])
    rows = payload.get("table", {}).get("rows", [])
    headers = [((c.get("label") or c.get("id") or "").strip()) for c in cols]
    # Locate NSE Code column, matching the Excel parser semantics.
    symbol_idx = next((i for i, h in enumerate(headers) if "nse code" in h.lower()), None)
    if symbol_idx is None:
        raise ValueError("Could not find an 'NSE Code' column in the Google Sheet.")
    car_idx = next((i for i, h in enumerate(headers) if h.strip().lower() == "car"), None)
    tickers, seen = [], set()
    for r in rows:
        cells = r.get("c", []) if isinstance(r, dict) else []
        def cell_value(i):
            if i >= len(cells) or cells[i] is None:
                return None
            return cells[i].get("v")
        raw = cell_value(symbol_idx)
        if raw is None or not str(raw).strip():
            continue
        code = str(raw).strip().upper()
        if code in ("NSE CODE", "STOCKS IN BULL RUN"):
            continue
        symbol = code if code.endswith(".NS") else f"{code}.NS"
        if symbol in seen:
            continue
        seen.add(symbol)
        car = cell_value(car_idx) if car_idx is not None else None
        tickers.append({"symbol": symbol, "name": code, "car": car})
    if not tickers:
        raise ValueError("The Google Sheet contains an NSE Code header but no ticker rows.")
    return {"sheet": preferred_sheet, "tickers": tickers, "count": len(tickers)}
